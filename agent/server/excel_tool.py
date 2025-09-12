"""
mcp_excel_toolkit.py
FastMCP server – Excel/CSV extraction → Markdown (no camel dependency)

File Reading Strategy:
All Excel/CSV files for processing should be placed in the unified workspace directory 
(task_cache_dir/workspace/) as managed by agent_core.py. This tool receives
the workspace directory path and filename as separate parameters, then combines
them for file operations to maintain consistency with the code execution environment.
"""

# --------------------------------------------------------------------------- #
#  Imports
# --------------------------------------------------------------------------- #
import os
from typing import Optional, Tuple
from pathlib import Path

import anyio
import pandas as pd
from loguru import logger
from mcp.server.fastmcp import FastMCP


# --------------------------------------------------------------------------- #
#  Unified path management functions
# --------------------------------------------------------------------------- #
def _get_workspace_dir(task_cache_dir: Optional[str] = None) -> str:
    """
    Get or create the unified workspace directory for a task.
    This follows the same pattern as code_tool.py to ensure consistency.
    
    📁 DIRECTORY STRUCTURE:
        task_cache_dir/workspace/  (flat structure - all files here)
    
    Args:
        task_cache_dir: Task-specific cache directory path
        
    Returns:
        str: Absolute path to workspace directory
    """
    if not task_cache_dir:
        task_cache_dir = os.getenv("AGENT_CACHE_DIR", "./workspaces")
    ospath = os.getenv("OSPATH")
    if ospath:
        workspace_dir = Path(ospath) / task_cache_dir / "workspace"
    else:
        workspace_dir = Path(task_cache_dir) / "workspace"
    workspace_dir.mkdir(parents=True, exist_ok=True)
    return str(workspace_dir)


def _build_file_path(workspace_dir: str, filename: str) -> str:
    """
    Build full file path from workspace directory and filename.
    Ensures all files are accessed from the unified workspace location.
    
    Args:
        workspace_dir: Base workspace directory path
        filename: Target filename (should not contain path separators)
        
    Returns:
        str: Complete file path for document processing
    """
    # Normalize filename to prevent directory traversal
    clean_filename = os.path.basename(filename)
    return str(Path(workspace_dir) / clean_filename)


# --------------------------------------------------------------------------- #
#  Helper class
# --------------------------------------------------------------------------- #
class ExcelToolkit:
    """
    Extracts rich information from Excel (.xls/.xlsx) or CSV files:
      • Every sheet converted to Markdown using `tabulate`
      • List of all cell coordinates with value + font / fill RGB colours
      
    **WORKSPACE INTEGRATION**: 
    All Excel/CSV files for processing must be located in the unified workspace directory
    (task_cache_dir/workspace/) as established by the code execution environment.
    This ensures consistent file access across all tools in the agent system.
    """

    def __init__(self, timeout: Optional[float] = None):
        self.timeout = timeout

    # ---------- public synchronous API ------------------------------------ #
    def extract_excel_content(self, workspace_dir: str, filename: str) -> str:
        """
        Extract content from Excel/CSV file located in the unified workspace.
        
        Args:
            workspace_dir: Path to the unified workspace directory
            filename: Name of the Excel/CSV file to process (within workspace)
            
        Returns:
            str: Extracted content as formatted text/Markdown
        """
        # Build full document path using unified workspace structure
        document_path = _build_file_path(workspace_dir, filename)
        
        if not filename.lower().endswith((".xls", ".xlsx", ".csv")):
            raise ValueError("Only .xls, .xlsx or .csv files are supported.")

        logger.info(f"Processing Excel/CSV file: workspace={workspace_dir}, filename={filename}, full_path={document_path}")

        # Verify file exists in workspace
        if not os.path.exists(document_path):
            raise FileNotFoundError(f"Excel/CSV file '{filename}' not found in workspace directory: {workspace_dir}")

        if document_path.lower().endswith(".csv"):
            return self._handle_csv(document_path)

        # If it's .xls, convert to .xlsx first (save converted file in workspace)
        if document_path.lower().endswith(".xls"):
            from xls2xlsx import XLS2XLSX

            xlsx_filename = os.path.splitext(filename)[0] + ".xlsx"
            out_path = _build_file_path(workspace_dir, xlsx_filename)
            XLS2XLSX(document_path).to_xlsx(out_path)
            document_path = out_path
            logger.debug(f"Converted .xls → .xlsx in workspace: {xlsx_filename}")

        return self._handle_xlsx(document_path, workspace_dir)

    # ---------- helpers ---------------------------------------------------- #
    def _handle_csv(self, path: str) -> str:
        """
        Process CSV file and return formatted content.
        
        Args:
            path: Full path to the CSV file
            
        Returns:
            str: Formatted CSV content as Markdown
        """
        try:
            df = pd.read_csv(path)
        except Exception as e:
            logger.error(f"CSV read failed: {e}")
            raise

        return "CSV File Processed:\n" + self._df_to_md(df)

    def _handle_xlsx(self, path: str, workspace_dir: str) -> str:
        """
        Process Excel (.xlsx) file and return formatted content.
        
        Args:
            path: Full path to the Excel file
            workspace_dir: Workspace directory for any temporary files
            
        Returns:
            str: Formatted Excel content as Markdown with cell information
        """
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        output_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cells_info = []

            for row in ws.iter_rows():
                for cell in row:
                    coord = f"{cell.row}{cell.column_letter}"
                    font_rgb = (
                        cell.font.color.rgb
                        if cell.font and cell.font.color and cell.font.color.rgb
                        else None
                    )
                    fill_rgb = (
                        cell.fill.fgColor.rgb
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb
                        else None
                    )
                    cells_info.append(
                        {
                            "index": coord,
                            "value": cell.value,
                            "font_color": font_rgb,
                            "fill_color": fill_rgb,
                        }
                    )

            # Re‑load sheet via pandas for prettier Markdown
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")

            part = (
                f"Sheet Name: {sheet_name}\n"
                f"Cell information list:\n{cells_info}\n\n"
                f"Markdown View of the content:\n{self._df_to_md(df)}\n"
                + "-" * 40
            )
            output_parts.append(part)

        return "\n".join(output_parts)

    @staticmethod
    def _df_to_md(df: pd.DataFrame) -> str:
        """
        Convert pandas DataFrame to Markdown table format.
        
        Args:
            df: Pandas DataFrame to convert
            
        Returns:
            str: DataFrame as Markdown table
        """
        from tabulate import tabulate

        return tabulate(df, headers="keys", tablefmt="pipe")


# --------------------------------------------------------------------------- #
#  FastMCP server
# --------------------------------------------------------------------------- #
mcp = FastMCP("excel_toolkit")
toolkit = ExcelToolkit()


@mcp.tool()
async def extract_excel_content(
    filename: str,
    task_cache_dir: str | None = None
) -> str:
    """
    Extract and return a Markdown‑rich description of an Excel/CSV file from the unified workspace.
    
    **File Location Requirements**:
    - Excel/CSV files must be placed in the workspace directory (task_cache_dir/workspace/)
    - This ensures consistency with code execution and other tools
    - The agent_core.py manages the task_cache_dir and passes it to tools
    
    Args:
        filename (str): Name of the Excel/CSV file to process (.xls, .xlsx, .csv)
        task_cache_dir (str, optional): Task-specific cache directory path.
                                       If not provided, uses AGENT_CACHE_DIR environment variable.
    
    Returns:
        str: Extracted Excel/CSV content formatted as Markdown with cell details
        
    Example Usage:
        await extract_excel_content("data.xlsx", "/path/to/task_cache")
        await extract_excel_content("report.csv")
    """
    # Get unified workspace directory
    workspace_dir = _get_workspace_dir(task_cache_dir)
    
    # Run the synchronous extractor in a worker thread so the event loop
    # stays free.
    try:
        return await anyio.to_thread.run_sync(
            toolkit.extract_excel_content, workspace_dir, filename
        )
    except Exception as e:
        logger.error(f"Excel processing failed for {filename}: {e}")
        raise ValueError(f"Excel processing error: {str(e)}")


# --------------------------------------------------------------------------- #
#  Entrypoint
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    mcp.run(transport="stdio")
